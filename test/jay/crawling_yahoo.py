from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from datetime import date, datetime
import os
import time
from tqdm import tqdm  # tqdm 라이브러리 추가

def setup_browser():
    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)
    chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    chrome_options.add_argument('--headless') 
    chrome_options.add_argument('--ignore-ssl-errors=yes')
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--log-level=3')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--incognito')

    # Add Image Loading inactive Flag to reduce loading time
    chrome_options.add_argument('--disable-images')
    chrome_options.add_experimental_option(
        "prefs", {'profile.managed_default_content_settings.images': 2})
    chrome_options.add_argument('--blink-settings=imagesEnabled=false')
    service = Service(executable_path=ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=chrome_options)

def scroll_page(browser, pause_time=0.2, scroll_count=5):
    for _ in range(scroll_count):
        browser.find_element(by=By.TAG_NAME, value="body").send_keys(Keys.PAGE_DOWN)
        time.sleep(pause_time)
        
def format_date(published_at):
    # 날짜 형식 변환
    try:
        if "PM" in published_at:
            original_date_str = published_at.split("PM")[0].strip() + " PM"
        elif "AM" in published_at:
            original_date_str = published_at.split("AM")[0].strip() + " AM"
        else:
            original_date_str = published_at

        original_date = datetime.strptime(original_date_str, "%a, %b %d, %Y, %I:%M %p")
        formatted_date = original_date.strftime("%Y.%m.%d. %p %I:%M")
        formatted_date = formatted_date.replace("AM", "오전").replace("PM", "오후")

        return formatted_date
    except ValueError:
        return published_at  # 변환 실패 시 원본 반환
    
def scrape_articles(browser):
    articles = browser.find_elements(by=By.XPATH, value="//ul/li[contains(@class, 'js-stream-content')]")
    data = []
    original_window = browser.current_window_handle  # 현재 창 핸들 저장
    
    for article in tqdm(articles, desc="Scraping Articles", unit="article"):  # tqdm 추가
        try:
            category = article.find_element(by=By.CSS_SELECTOR, value="[data-test-locator='catlabel']").text
            title = article.find_element(by=By.TAG_NAME, value="h3").text
            image_tag = article.find_element(by=By.TAG_NAME, value="img")
            image = image_tag.get_attribute("src")
            
            link = article.find_element(by=By.TAG_NAME, value="a").get_attribute("href")
            newspaper = article.find_element(by=By.CSS_SELECTOR, value="span").text

            # 새 탭 열기
            browser.execute_script("window.open(arguments[0]);", link)
            time.sleep(2)  # 페이지 로드 대기
            browser.switch_to.window(browser.window_handles[1])  # 새 탭으로 전환

            # 본문 텍스트 수집
            caas_body_element = browser.find_element(by=By.CSS_SELECTOR, value=".caas-body")
            main = caas_body_element.text

            # published_at 가져오기
            try:
                published_at = browser.find_element(by=By.TAG_NAME, value="time").text
                if published_at:
                    published_at = format_date(published_at)
            except:
                published_at = ""
            
            stock_element = browser.find_element(by=By.CSS_SELECTOR, value=".caas-xray-entity a")
            stock_name = stock_element.text
            stock_url=stock_element.get_attribute("href")
            
            # 데이터 저장
            data.append([newspaper, category, title, main, published_at, link, image, stock_name, stock_url])
            
            # 새 탭 닫기
            browser.close()
            browser.switch_to.window(original_window)  # 원래 창으로 돌아오기
            
        except Exception as e:
            print(f"Error occurred: {e}")
            continue
            
    return data

def save_to_excel(data):
    today = date.today().strftime("%Y-%m-%d")
    header = ["source", "category", "title", "main", "published_at", "link", "image", "stock_name", "stock_url"]

    if os.path.exists("크롤링 데이터.xlsx"):
        write_wb = load_workbook("크롤링 데이터.xlsx")
        write_ws = write_wb.create_sheet(today)
    else:
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.title = today
        write_ws.append(header)

    for item in data:
        write_ws.append(item)
    write_wb.save("크롤링 데이터.xlsx")

def main():
    browser = setup_browser()
    browser.get("https://finance.yahoo.com/topic/stock-market-news/")
    scroll_page(browser)
    data = scrape_articles(browser)
    save_to_excel(data)
    browser.quit()

if __name__ == "__main__":
    main()
