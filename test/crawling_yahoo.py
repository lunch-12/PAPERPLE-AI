from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from datetime import date, datetime
import os
import time

def setup_browser():
    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)
    chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    chrome_options.add_argument('--headless') 
    chrome_options.add_argument('--ignore-ssl-errors=yes')
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--log-level=3')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--incognito')

    # Add Image Loading inactive Flag to reduce loading time
    chrome_options.add_argument('--disable-images')
    chrome_options.add_experimental_option(
        "prefs", {'profile.managed_default_content_settings.images': 2})
    chrome_options.add_argument('--blink-settings=imagesEnabled=false')
    service = Service(executable_path=ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=chrome_options)

def scroll_page(browser, pause_time=0.2, scroll_count=50):
    for _ in range(scroll_count):
        browser.find_element(by=By.TAG_NAME, value="body").send_keys(Keys.PAGE_DOWN)
        time.sleep(pause_time)
        
def format_date(published_at):
    # 날짜 형식 변환
    try:
        # 'PM' 또는 'AM' 이후의 시간대 정보를 제거
        if "PM" in published_at:
            original_date_str = published_at.split("PM")[0].strip() + " PM"
        elif "AM" in published_at:
            original_date_str = published_at.split("AM")[0].strip() + " AM"
        else:
            original_date_str = published_at

        # 원래 문자열을 datetime 객체로 변환
        original_date = datetime.strptime(original_date_str, "%a, %b %d, %Y, %I:%M %p")

        # 원하는 형식으로 변환하여 출력 
        formatted_date = original_date.strftime("%Y.%m.%d. %p %I:%M")

        # 오전/오후 형식을 한글로 변경
        formatted_date = formatted_date.replace("AM", "오전").replace("PM", "오후")

        return formatted_date
    except ValueError:
        return published_at  # 변환 실패 시 원본 반환
    
def scrape_articles(browser):
    articles = browser.find_elements(by=By.XPATH, value="//ul/li[contains(@class, 'js-stream-content')]")
    data = []
    original_window = browser.current_window_handle  # 현재 창 핸들 저장
    
    for article in articles:
        try:
            category = article.find_element(by=By.CSS_SELECTOR, value="[data-test-locator='catlabel']").text
            title = article.find_element(by=By.TAG_NAME, value="h3").text
            image_tag = article.find_element(by=By.TAG_NAME, value="img")
            image = image_tag.get_attribute("src")
            
            link = article.find_element(by=By.TAG_NAME, value="a").get_attribute("href")
            newspaper = article.find_element(by=By.CSS_SELECTOR, value="span").text

            # 새 탭 열기
            browser.execute_script("window.open(arguments[0]);", link)
            time.sleep(2)  # 페이지 로드 대기
            browser.switch_to.window(browser.window_handles[1])  # 새 탭으로 전환

            # 본문 텍스트 수집
            caas_body_element = browser.find_element(by=By.CSS_SELECTOR, value=".caas-body")
            main = caas_body_element.text
            
            # published_at 가져오기
            try:
                published_at = browser.find_element(by=By.TAG_NAME, value="time").text
                if published_at:
                    published_at = format_date(published_at)
            except:
                published_at = ""
            
            # 관련 종목 수집
            try:
                stock_element = browser.find_element(by=By.CSS_SELECTOR, value=".caas-xray-entity a")
                stock_name = stock_element.text
                stock_url = stock_element.get_attribute("href")
            except:
                # 요소가 존재하지 않으면 빈 문자열 할당
                stock_name = ""
                stock_url = ""


            # 데이터 저장
            data.append([newspaper, category, title, main, published_at, link, image, stock_name, stock_url])
            
            # 새 탭 닫기
            browser.close()
            browser.switch_to.window(original_window)  # 원래 창으로 돌아오기
            
        except Exception as e:
            print(f"Error occurred: {e}")
            continue
            
    return data

def save_to_excel(data):
    # 현재 날짜와 시간 가져오기
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"./data/yahoo_article.xlsx"
    
    # 엑셀 파일에 저장할 데이터의 헤더 정의
    header = ["source", "category", "title", "main", "published_at", "link", "image", "stock_name", "stock_url"]

    # 엑셀 파일이 존재하는 경우 해당 파일에 새로운 시트를 추가, 존재하지 않으면 새 파일 생성
    if os.path.exists(file_name):
        write_wb = load_workbook(file_name)
        write_ws = write_wb.create_sheet(now)
    else:
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.title = now
        write_ws.append(header)

    # 데이터를 엑셀 파일에 추가
    for item in data:
        write_ws.append(item)
    
    # 파일 저장
    write_wb.save(file_name)

def main():
    browser = setup_browser()
    browser.get("https://finance.yahoo.com/topic/stock-market-news/")
    scroll_page(browser)
    data = scrape_articles(browser)
    save_to_excel(data)
    browser.quit()

if __name__ == "__main__":
    main()